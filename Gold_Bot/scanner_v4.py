"""
scanner_v4.py — Manny's Gold Strategy V4 — Tier 1 Upgrades Edition
====================================================================
Purpose:
    Connects to MetaTrader 5 (MT5) and scans multiple symbols across multiple
    timeframes for high-probability Smart Money Concepts (SMC) trade signals.
    Signals are filtered by bias alignment, structure, order blocks, FVGs,
    and a confidence score gate before auto-executing or alerting manually.

    NEW IN V4 (Tier 1 Upgrades):
    ✅ T1-1: Partial Profit Engine — closes 50% of position at 1:1 RR
    ✅ T1-2: Break-Even Stop Move — moves SL to entry after partial close
    ✅ T1-3: Trailing Stop on Runners — trails SL behind swing points
    ✅ T1-4: Dynamic Exit on Reversal — exits early if reversal confirmed while in profit
    ✅ T1-5: Early SL Exit — exits early if no reversal seen while in drawdown
    ✅ T1-6: Daily Drawdown Limit — shuts bot down after 3% daily loss
    ✅ T1-7: Enhanced Excel Trade Journal — colour-coded, detailed columns

Inputs:
    - MT5 demo/live account credentials (via .env file)
    - Telegram bot token and chat ID (via .env file)
    - Live OHLCV data pulled from MT5 via MetaTrader5 Python API

Outputs:
    - Auto-executed trades on MT5 for signals scoring >= session threshold
    - Telegram alerts for manual review on signals scoring 5-6
    - trade_journal.xlsx: colour-coded Excel audit trail of every signal

Author: Emmanuel Ogbu (Manny)
Date: May 2026
"""

import os                           # built-in: file system checks and env vars
import time                         # built-in: sleep between scan loops
import requests                     # third-party: HTTP requests for Telegram API
import MetaTrader5 as mt5           # MT5 Python API for live trading connection
import pandas as pd                 # data manipulation and OHLCV DataFrame handling
import numpy as np                  # numerical operations
from dotenv import load_dotenv      # loads environment variables from .env file
from datetime import datetime, timezone, timedelta  # timezone-aware datetime handling
from openpyxl import load_workbook, Workbook        # Excel file creation and editing
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side  # Excel styling

load_dotenv()  # load credentials from .env file into os.environ

# ═══════════════════════════════════════════
# CREDENTIALS — loaded from .env file
# Never hardcode credentials directly in code
# ═══════════════════════════════════════════
MT5_LOGIN        = int(os.getenv("MT5_LOGIN"))
MT5_PASSWORD     = os.getenv("MT5_PASSWORD")
MT5_SERVER       = os.getenv("MT5_SERVER")
TELEGRAM_TOKEN   = os.getenv("TELEGRAM_TOKEN")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")

# ═══════════════════════════════════════════
# SYMBOLS TO SCAN
# ═══════════════════════════════════════════
SYMBOLS       = ["XAUUSD", "XAGUSD", "US500", "US30", "BTCUSD", "USTEC", "NZDUSD"]
THREE_AM_PAIRS = {"XAGUSD", "US500", "US30", "USOIL"}

# ═══════════════════════════════════════════
# TIMEFRAMES PER SYMBOL
# ═══════════════════════════════════════════
SYMBOL_TIMEFRAMES = {
    "XAUUSD":  [mt5.TIMEFRAME_M5, mt5.TIMEFRAME_M15, mt5.TIMEFRAME_M30, mt5.TIMEFRAME_H1, mt5.TIMEFRAME_H4],
    "XAGUSD":  [mt5.TIMEFRAME_M5, mt5.TIMEFRAME_M15, mt5.TIMEFRAME_M30, mt5.TIMEFRAME_H1, mt5.TIMEFRAME_H4],
    "US500":   [mt5.TIMEFRAME_M5, mt5.TIMEFRAME_M15, mt5.TIMEFRAME_M30, mt5.TIMEFRAME_H1, mt5.TIMEFRAME_H4],
    "US30":    [mt5.TIMEFRAME_M5, mt5.TIMEFRAME_M15, mt5.TIMEFRAME_M30, mt5.TIMEFRAME_H1, mt5.TIMEFRAME_H4],
    "BTCUSD":  [mt5.TIMEFRAME_M5, mt5.TIMEFRAME_M15, mt5.TIMEFRAME_M30, mt5.TIMEFRAME_H1, mt5.TIMEFRAME_H4],
    "USTEC":   [mt5.TIMEFRAME_M5, mt5.TIMEFRAME_M15, mt5.TIMEFRAME_M30, mt5.TIMEFRAME_H1, mt5.TIMEFRAME_H4],
    "NZDUSD":  [mt5.TIMEFRAME_M15, mt5.TIMEFRAME_M30, mt5.TIMEFRAME_H1, mt5.TIMEFRAME_H4],
}

TIMEFRAME_NAMES = {
    mt5.TIMEFRAME_M5:  "5M",
    mt5.TIMEFRAME_M15: "15M",
    mt5.TIMEFRAME_M30: "30M",
    mt5.TIMEFRAME_H1:  "1H",
    mt5.TIMEFRAME_H4:  "4H",
}

# ═══════════════════════════════════════════
# GLOBAL SETTINGS
# ═══════════════════════════════════════════
SYMBOL         = "XAUUSD"
TIMEFRAME_4H   = mt5.TIMEFRAME_H4
TIMEFRAME_1H   = mt5.TIMEFRAME_H1
RISK_PERCENT   = 0.01        # 1% account risk per trade
RR             = 3.0         # 1:3 reward-to-risk ratio
CHECK_INTERVAL = 60          # seconds between scan loops
ADX_PERIOD     = 14
ADX_THRESH     = 25
ATR_PERIOD     = 14
SCALP_MIN_RR   = 1.5         # minimum RR for scalp signals

# ═══════════════════════════════════════════
# SESSION HOURS (UTC)
# ═══════════════════════════════════════════
LONDON_START  = 7
LONDON_END    = 12
OVERLAP_START = 12
OVERLAP_END   = 13
NY_START      = 13
NY_END        = 17
POSTNY_START  = 17
POSTNY_END    = 22

# ═══════════════════════════════════════════
# CONFIDENCE SCORE THRESHOLDS
# ═══════════════════════════════════════════
CONFIDENCE_THRESHOLD_PRIME  = 7
CONFIDENCE_THRESHOLD_POSTNY = 8
CONFIDENCE_MANUAL_MIN       = 5

# ═══════════════════════════════════════════
# T1-6: DAILY DRAWDOWN LIMIT
# If total daily losses exceed this % of account, bot stops for the day.
# Resets at midnight UTC automatically.
# ─────────────────────────────────────────
DAILY_DRAWDOWN_LIMIT = 0.03   # 3% of account balance = max daily loss allowed

# ═══════════════════════════════════════════
# T1-1/T1-2: PARTIAL PROFIT SETTINGS
# PARTIAL_CLOSE_RR: RR level at which we close 50% and move SL to BE
# PARTIAL_CLOSE_PCT: percentage of position to close at that level
# ─────────────────────────────────────────
PARTIAL_CLOSE_RR  = 1.0    # close partial when trade hits 1:1 RR
PARTIAL_CLOSE_PCT = 0.5    # close 50% of the position

# ═══════════════════════════════════════════
# T1-3: TRAILING STOP SETTINGS
# Trail stop behind swing high/low using a lookback window
# ─────────────────────────────────────────
TRAIL_LOOKBACK = 5   # number of bars to look back for trailing swing point

# ═══════════════════════════════════════════
# T1-7: EXCEL TRADE JOURNAL FILE
# ─────────────────────────────────────────
TRADE_JOURNAL_FILE = "trade_journal.xlsx"

# All columns in the Excel journal — much more detailed than old CSV
JOURNAL_HEADERS = [
    "timestamp",          # UTC time signal fired
    "symbol",             # e.g. XAUUSD
    "timeframe",          # e.g. 30M
    "tier",               # STRONG / MEDIUM / SCALP
    "confidence_score",   # 0-10 at time of signal
    "entry",              # entry price
    "sl",                 # stop loss price
    "tp",                 # take profit price
    "trigger_type",       # OB / FVG / Sweep / Pin / PDH / CHoCH
    "sl_source",          # OB / FVG / Swing / ATR
    "session",            # London / Overlap / New York / Post-NY
    "status",             # auto-executed / manual-alert / ignored
    "partial_taken",      # YES / NO — did partial profit fire?
    "partial_price",      # price at which partial was taken
    "be_moved",           # YES / NO — was SL moved to break-even?
    "exit_reason",        # TP / SL / Early-Profit / Early-SL / Trailing
    "exit_price",         # actual price trade closed at
    "actual_pnl",         # actual profit or loss in account currency
    "time_in_trade_mins", # how long trade was open in minutes
    "rr_achieved",        # actual RR achieved (e.g. 1.8 if closed early)
]

# ═══════════════════════════════════════════
# BIAS CANDLE HOURS (UTC)
# ═══════════════════════════════════════════
BIAS_CANDLE1_HOUR = 0   # 00:00 UTC candle 1 for all pairs
BIAS_CANDLE2_HOUR = 4   # 04:00 UTC candle 2 for all pairs

# ═══════════════════════════════════════════
# NEWS BLACKOUT DATES
# ═══════════════════════════════════════════
CPI_DATES = [
    datetime(2024,1,11,13,30,tzinfo=timezone.utc), datetime(2024,2,13,13,30,tzinfo=timezone.utc),
    datetime(2024,3,12,13,30,tzinfo=timezone.utc), datetime(2024,4,10,13,30,tzinfo=timezone.utc),
    datetime(2024,5,15,13,30,tzinfo=timezone.utc), datetime(2024,6,12,13,30,tzinfo=timezone.utc),
    datetime(2024,7,11,13,30,tzinfo=timezone.utc), datetime(2024,8,14,13,30,tzinfo=timezone.utc),
    datetime(2024,9,11,13,30,tzinfo=timezone.utc), datetime(2024,10,10,13,30,tzinfo=timezone.utc),
    datetime(2024,11,13,13,30,tzinfo=timezone.utc),datetime(2024,12,11,13,30,tzinfo=timezone.utc),
    datetime(2025,1,15,13,30,tzinfo=timezone.utc), datetime(2025,2,12,13,30,tzinfo=timezone.utc),
    datetime(2025,3,12,13,30,tzinfo=timezone.utc), datetime(2025,4,9,13,30,tzinfo=timezone.utc),
    datetime(2025,5,13,13,30,tzinfo=timezone.utc), datetime(2025,6,11,13,30,tzinfo=timezone.utc),
    datetime(2025,7,11,13,30,tzinfo=timezone.utc), datetime(2025,8,12,13,30,tzinfo=timezone.utc),
    datetime(2025,9,10,13,30,tzinfo=timezone.utc), datetime(2025,10,9,13,30,tzinfo=timezone.utc),
    datetime(2025,11,12,13,30,tzinfo=timezone.utc),datetime(2025,12,10,13,30,tzinfo=timezone.utc),
    datetime(2026,1,14,13,30,tzinfo=timezone.utc), datetime(2026,2,12,13,30,tzinfo=timezone.utc),
    datetime(2026,3,11,13,30,tzinfo=timezone.utc), datetime(2026,4,10,13,30,tzinfo=timezone.utc),
    datetime(2026,5,13,13,30,tzinfo=timezone.utc), datetime(2026,6,11,13,30,tzinfo=timezone.utc),
    datetime(2026,7,14,13,30,tzinfo=timezone.utc), datetime(2026,8,12,13,30,tzinfo=timezone.utc),
    datetime(2026,9,10,13,30,tzinfo=timezone.utc), datetime(2026,10,8,13,30,tzinfo=timezone.utc),
    datetime(2026,11,12,13,30,tzinfo=timezone.utc),datetime(2026,12,9,13,30,tzinfo=timezone.utc),
]

FOMC_DATES = [
    datetime(2024,1,31,19,0,tzinfo=timezone.utc),  datetime(2024,3,20,19,0,tzinfo=timezone.utc),
    datetime(2024,5,1,19,0,tzinfo=timezone.utc),   datetime(2024,6,12,19,0,tzinfo=timezone.utc),
    datetime(2024,7,31,19,0,tzinfo=timezone.utc),  datetime(2024,9,18,19,0,tzinfo=timezone.utc),
    datetime(2024,11,7,19,0,tzinfo=timezone.utc),  datetime(2024,12,18,19,0,tzinfo=timezone.utc),
    datetime(2025,1,29,19,0,tzinfo=timezone.utc),  datetime(2025,3,19,19,0,tzinfo=timezone.utc),
    datetime(2025,5,7,19,0,tzinfo=timezone.utc),   datetime(2025,6,18,19,0,tzinfo=timezone.utc),
    datetime(2025,7,30,19,0,tzinfo=timezone.utc),  datetime(2025,9,17,19,0,tzinfo=timezone.utc),
    datetime(2025,11,5,19,0,tzinfo=timezone.utc),  datetime(2025,12,17,19,0,tzinfo=timezone.utc),
    datetime(2026,1,28,19,0,tzinfo=timezone.utc),  datetime(2026,3,18,19,0,tzinfo=timezone.utc),
    datetime(2026,5,6,19,0,tzinfo=timezone.utc),   datetime(2026,6,17,19,0,tzinfo=timezone.utc),
    datetime(2026,7,29,19,0,tzinfo=timezone.utc),  datetime(2026,9,16,19,0,tzinfo=timezone.utc),
    datetime(2026,11,4,19,0,tzinfo=timezone.utc),  datetime(2026,12,16,19,0,tzinfo=timezone.utc),
]

NEWS_WINDOW = timedelta(minutes=30)


# ═══════════════════════════════════════════
# T1-6: DAILY DRAWDOWN TRACKER
# ─────────────────────────────────────────────
# Tracks daily P&L and shuts bot down if limit is breached.
# Resets automatically at midnight UTC each day.
# Uses a simple dict so state persists across the scan loop.
# ═══════════════════════════════════════════
daily_pnl_state = {
    "date":       None,    # current UTC date (resets tracking when date changes)
    "start_bal":  None,    # account balance at start of trading day
    "current_pnl": 0.0,   # running P&L for today
    "shutdown":   False,   # True = bot is shut down for today due to drawdown limit
}


def check_daily_drawdown() -> bool:
    """
    T1-6: Check whether the daily drawdown limit has been breached.
    Resets state at midnight UTC each day.

    Returns:
        True if bot should STOP trading (limit breached), False if safe to continue.
    """
    global daily_pnl_state

    now      = datetime.now(timezone.utc)
    today    = now.date()
    account  = mt5.account_info()

    if account is None:
        return False   # can't check, allow trading to continue

    # ── Reset at midnight UTC (new trading day) ──
    if daily_pnl_state["date"] != today:
        daily_pnl_state["date"]       = today
        daily_pnl_state["start_bal"]  = account.balance   # record balance at day start
        daily_pnl_state["current_pnl"] = 0.0
        daily_pnl_state["shutdown"]   = False              # lift yesterday's shutdown
        print(f"📅 New trading day — Balance reset: {account.balance:.2f} | Drawdown limit: {account.balance * DAILY_DRAWDOWN_LIMIT:.2f}")

    # ── Calculate today's P&L ──
    # equity = balance + floating P&L (open trades included)
    # We compare equity to start balance to get true daily P&L including open trades
    daily_pnl = account.equity - daily_pnl_state["start_bal"]
    daily_pnl_state["current_pnl"] = daily_pnl

    # ── Check if limit breached ──
    max_loss = daily_pnl_state["start_bal"] * DAILY_DRAWDOWN_LIMIT   # e.g. 1500 on 50k account

    if daily_pnl <= -max_loss and not daily_pnl_state["shutdown"]:
        # First time breaching today — trigger shutdown and alert
        daily_pnl_state["shutdown"] = True
        msg = (
            f"🛑 <b>DAILY DRAWDOWN LIMIT HIT — BOT SHUTDOWN</b>\n\n"
            f"📉 Today's Loss: {daily_pnl:.2f}\n"
            f"🚫 Limit: -{max_loss:.2f} ({DAILY_DRAWDOWN_LIMIT*100:.0f}% of {daily_pnl_state['start_bal']:.2f})\n"
            f"⏰ Shutting down until midnight UTC\n"
            f"✅ Bot will restart automatically tomorrow"
        )
        send_telegram(msg)
        print(f"🛑 DAILY DRAWDOWN LIMIT REACHED: {daily_pnl:.2f} loss. Bot shutdown.")

    return daily_pnl_state["shutdown"]   # True = stop trading, False = continue


# ═══════════════════════════════════════════
# T1-7: ENHANCED EXCEL TRADE JOURNAL
# ─────────────────────────────────────────────
# Creates a colour-coded Excel file instead of CSV.
# Green rows = profitable trades / partial profits taken
# Red rows = losses / stopped out trades
# Yellow rows = manual alerts (not auto-executed)
# Grey rows = ignored signals
# Columns are bolded headers with borders for readability.
# ═══════════════════════════════════════════

# Define colour fills for each trade outcome
FILL_GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # profit
FILL_RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # loss
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # manual
FILL_GREY   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # ignored
FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # header bg

FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)   # white bold text for header row
FONT_NORMAL = Font(size=10)

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)


def init_journal() -> None:
    """
    T1-7: Create the Excel journal file with formatted headers if it doesn't exist.
    Called once at bot startup. Safe to call multiple times — skips if file exists.
    """
    if os.path.isfile(TRADE_JOURNAL_FILE):
        return   # file already exists, don't overwrite existing data

    # Create new workbook and set up the sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Journal"

    # Write header row with styling
    for col_idx, header in enumerate(JOURNAL_HEADERS, start=1):
        cell            = ws.cell(row=1, column=col_idx, value=header)
        cell.fill       = FILL_HEADER    # dark blue background
        cell.font       = FONT_HEADER    # white bold text
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = THIN_BORDER

    # Set column widths for readability
    col_widths = [20, 8, 8, 8, 10, 10, 10, 10, 12, 10, 10, 14, 12, 12, 10, 14, 10, 12, 16, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.row_dimensions[1].height = 30   # taller header row

    wb.save(TRADE_JOURNAL_FILE)
    print(f"📊 trade_journal.xlsx created — colour-coded Excel journal ready")


def log_trade_excel(
    symbol: str,
    timeframe: str,
    tier: str,
    confidence_score: int,
    entry_price: float,
    sl_price: float,
    tp_price: float,
    trigger_type: str,
    sl_source: str,
    session: str,
    status: str,
    partial_taken: str = "NO",
    partial_price: float = None,
    be_moved: str = "NO",
    exit_reason: str = "",
    exit_price: float = None,
    actual_pnl: float = None,
    time_in_trade_mins: float = None,
    rr_achieved: float = None,
) -> None:
    """
    T1-7: Append a trade record to the Excel journal with colour coding.

    Colour logic:
        Green  = actual_pnl > 0 (profitable trade)
        Red    = actual_pnl < 0 (losing trade)
        Yellow = status == 'manual-alert' (not auto-executed)
        Grey   = status == 'ignored'
        White  = status == 'auto-executed' but outcome not yet known

    Args:
        symbol:             Trading symbol e.g. 'XAUUSD'
        timeframe:          Timeframe string e.g. '30M'
        tier:               'STRONG', 'MEDIUM', or 'SCALP'
        confidence_score:   Integer 0-10
        entry_price:        Entry price at signal time
        sl_price:           Stop loss price
        tp_price:           Take profit price
        trigger_type:       What triggered the signal
        sl_source:          Where SL was anchored
        session:            Active session name
        status:             'auto-executed', 'manual-alert', or 'ignored'
        partial_taken:      'YES' or 'NO'
        partial_price:      Price at which partial was taken (or None)
        be_moved:           'YES' or 'NO' — was SL moved to break-even
        exit_reason:        How the trade closed e.g. 'TP', 'SL', 'Early-Profit'
        exit_price:         Actual close price (or None if still open)
        actual_pnl:         Actual P&L in account currency (or None if still open)
        time_in_trade_mins: Minutes trade was open (or None)
        rr_achieved:        Actual RR achieved (or None)
    """
    # Load existing workbook (or create if somehow missing)
    if not os.path.isfile(TRADE_JOURNAL_FILE):
        init_journal()

    wb = load_workbook(TRADE_JOURNAL_FILE)
    ws = wb.active

    # Find the next empty row by scanning down from row 2
    next_row = ws.max_row + 1

    # Build row data in same order as JOURNAL_HEADERS
    row_data = [
        datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S'),
        symbol,
        timeframe,
        tier,
        confidence_score,
        round(entry_price, 5),
        round(sl_price, 5),
        round(tp_price, 5),
        trigger_type,
        sl_source,
        session,
        status,
        partial_taken,
        round(partial_price, 5) if partial_price else "",
        be_moved,
        exit_reason,
        round(exit_price, 5) if exit_price else "",
        round(actual_pnl, 2) if actual_pnl is not None else "",
        round(time_in_trade_mins, 1) if time_in_trade_mins is not None else "",
        round(rr_achieved, 2) if rr_achieved is not None else "",
    ]

    # ── Determine row colour based on outcome ──
    if actual_pnl is not None and actual_pnl > 0:
        row_fill = FILL_GREEN    # profitable — green
    elif actual_pnl is not None and actual_pnl < 0:
        row_fill = FILL_RED      # loss — red
    elif status == "manual-alert":
        row_fill = FILL_YELLOW   # manual alert — yellow
    elif status == "ignored":
        row_fill = FILL_GREY     # ignored — grey
    else:
        row_fill = None          # auto-executed, outcome unknown yet — white

    # Write each cell in the row
    for col_idx, value in enumerate(row_data, start=1):
        cell           = ws.cell(row=next_row, column=col_idx, value=value)
        cell.font      = FONT_NORMAL
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if row_fill:
            cell.fill = row_fill   # apply colour if outcome known

    wb.save(TRADE_JOURNAL_FILE)


def update_trade_in_journal(
    symbol: str,
    timeframe: str,
    entry_price: float,
    partial_taken: str = None,
    partial_price: float = None,
    be_moved: str = None,
    exit_reason: str = None,
    exit_price: float = None,
    actual_pnl: float = None,
    time_in_trade_mins: float = None,
    rr_achieved: float = None,
) -> None:
    """
    T1-7: Update an existing trade row in the Excel journal when trade closes
    or when partial profit / break-even events occur.

    Finds the row by matching symbol + timeframe + entry_price.
    Updates only the columns that are provided (non-None).

    Args:
        symbol:             Symbol to find in journal
        timeframe:          Timeframe to find in journal
        entry_price:        Entry price to match (used to find the right row)
        partial_taken:      Update to 'YES' if partial was taken
        partial_price:      Price partial was taken at
        be_moved:           Update to 'YES' if SL moved to break-even
        exit_reason:        How trade closed
        exit_price:         Actual close price
        actual_pnl:         Actual P&L in account currency
        time_in_trade_mins: Minutes trade was open
        rr_achieved:        Actual RR achieved
    """
    if not os.path.isfile(TRADE_JOURNAL_FILE):
        return   # nothing to update if file doesn't exist

    wb = load_workbook(TRADE_JOURNAL_FILE)
    ws = wb.active

    # Column index mapping — must match JOURNAL_HEADERS order
    col_map = {header: idx + 1 for idx, header in enumerate(JOURNAL_HEADERS)}

    # Find the matching row by symbol + timeframe + entry price
    target_row = None
    for row in ws.iter_rows(min_row=2):
        row_sym   = row[col_map["symbol"] - 1].value
        row_tf    = row[col_map["timeframe"] - 1].value
        row_entry = row[col_map["entry"] - 1].value
        if (row_sym == symbol and row_tf == timeframe and
                row_entry is not None and abs(float(row_entry) - entry_price) < 0.001):
            target_row = row
            break   # found the matching trade row

    if target_row is None:
        print(f"  ⚠ Could not find trade in journal: {symbol} {timeframe} entry:{entry_price}")
        wb.save(TRADE_JOURNAL_FILE)
        return

    # Update only the fields that were provided
    def update_cell(col_name: str, value) -> None:
        """Helper: update a specific cell in the found row."""
        if value is not None:
            col_idx = col_map[col_name] - 1
            target_row[col_idx].value = value

    if partial_taken:     update_cell("partial_taken",      partial_taken)
    if partial_price:     update_cell("partial_price",      round(partial_price, 5))
    if be_moved:          update_cell("be_moved",           be_moved)
    if exit_reason:       update_cell("exit_reason",        exit_reason)
    if exit_price:        update_cell("exit_price",         round(exit_price, 5))
    if time_in_trade_mins: update_cell("time_in_trade_mins", round(time_in_trade_mins, 1))
    if rr_achieved:       update_cell("rr_achieved",        round(rr_achieved, 2))

    if actual_pnl is not None:
        update_cell("actual_pnl", round(actual_pnl, 2))
        # Also update row colour based on P&L outcome
        fill = FILL_GREEN if actual_pnl > 0 else FILL_RED
        for cell in target_row:
            if cell.row > 1:   # don't touch header row
                cell.fill = fill

    wb.save(TRADE_JOURNAL_FILE)


# ═══════════════════════════════════════════
# T1-1 / T1-2 / T1-3 / T1-4 / T1-5:
# OPEN TRADE MANAGEMENT
# ─────────────────────────────────────────────
# This is the core new function in V4.
# Runs every scan loop alongside signal detection.
# For every open position, it:
#   1. Checks if partial profit should be taken (1:1 RR reached)
#   2. Moves SL to break-even after partial
#   3. Trails SL behind new swing points
#   4. Detects reversal signals and exits early to lock profit
#   5. Detects aggressive drawdown with no reversal and exits early to limit loss
# ═══════════════════════════════════════════

# Track which trades have already had partial taken and BE moved
# Key = ticket number, Value = dict with management state
trade_management_state = {}


def get_reversal_signals(symbol: str, direction: str) -> bool:
    """
    T1-4 / T1-5: Detect whether reversal signals are present against an open trade.

    Checks for:
    - EMA crossover against trade direction
    - Strong opposing candle body (> 1.5x ATR)
    - Break of most recent swing point against direction

    Args:
        symbol:    Trading symbol e.g. 'XAUUSD'
        direction: 'BUY' or 'SELL' — direction of the open trade

    Returns:
        True if reversal signals are confirmed, False if not.
    """
    # Fetch the most recent 30M candles for reversal check
    # 30M is a good balance — not too noisy (5M), not too slow (4H)
    df = get_candles(mt5.TIMEFRAME_M30, 50, symbol)
    if df is None or len(df) < 20:
        return False   # no data — assume no reversal to be safe

    close  = df['close']
    open_  = df['open']
    atr    = calc_atr(df, 14).iloc[-2]
    ema50  = calc_ema(close, 50).iloc[-2]
    ema200 = calc_ema(close, 200).iloc[-2]
    last   = df.iloc[-2]   # last fully closed candle

    # ── Signal 1: EMA crossover against trade ──
    # If price crossed below EMA50 on a BUY trade, that's a warning signal
    prev_close = close.iloc[-3]
    if direction == "BUY" and prev_close > ema50 and last['close'] < ema50:
        return True   # price just crossed below EMA50 — BUY reversal signal
    if direction == "SELL" and prev_close < ema50 and last['close'] > ema50:
        return True   # price just crossed above EMA50 — SELL reversal signal

    # ── Signal 2: Strong opposing candle ──
    body = abs(last['close'] - last['open'])
    if direction == "BUY" and last['close'] < last['open'] and body > atr * 1.5:
        return True   # strong bearish candle against a BUY — reversal warning
    if direction == "SELL" and last['close'] > last['open'] and body > atr * 1.5:
        return True   # strong bullish candle against a SELL — reversal warning

    # ── Signal 3: Swing point break against direction ──
    swing_high = df['high'].iloc[-8:-2].max()   # highest high of last 6 closed bars
    swing_low  = df['low'].iloc[-8:-2].min()    # lowest low of last 6 closed bars

    if direction == "BUY" and last['close'] < swing_low:
        return True   # price broke below recent swing low — BUY structure broken
    if direction == "SELL" and last['close'] > swing_high:
        return True   # price broke above recent swing high — SELL structure broken

    return False   # no reversal signals detected


def get_trailing_sl(symbol: str, direction: str, current_sl: float) -> float:
    """
    T1-3: Calculate the new trailing stop loss based on recent swing points.

    For BUY trades: trail SL below the most recent swing low.
    For SELL trades: trail SL above the most recent swing high.
    Only moves SL in the favourable direction — never widens it.

    Args:
        symbol:     Trading symbol
        direction:  'BUY' or 'SELL'
        current_sl: Current stop loss price

    Returns:
        New SL price. If trail would widen SL, returns current_sl unchanged.
    """
    df = get_candles(mt5.TIMEFRAME_M30, 30, symbol)
    if df is None or len(df) < TRAIL_LOOKBACK + 3:
        return current_sl   # no data — keep current SL

    # Use recent swing points as trail reference
    recent_low  = df['low'].iloc[-(TRAIL_LOOKBACK + 2):-2].min()   # lowest low in lookback window
    recent_high = df['high'].iloc[-(TRAIL_LOOKBACK + 2):-2].max()  # highest high in lookback window

    atr = calc_atr(df, 14).iloc[-2]

    if direction == "BUY":
        # Trail SL just below the recent swing low (with small ATR buffer)
        new_sl = recent_low - atr * 0.2
        # Only update if new SL is HIGHER than current SL (trail moves up only)
        return new_sl if new_sl > current_sl else current_sl

    else:  # SELL
        # Trail SL just above the recent swing high (with small ATR buffer)
        new_sl = recent_high + atr * 0.2
        # Only update if new SL is LOWER than current SL (trail moves down only)
        return new_sl if new_sl < current_sl else current_sl


def manage_open_trades() -> None:
    """
    T1-1 / T1-2 / T1-3 / T1-4 / T1-5: Manage all currently open positions.

    Called every scan loop (every 60 seconds).
    For each open position placed by this bot (magic number 234000):

    Step 1 — Check partial profit: if trade reached 1:1 RR, close 50% and move SL to BE
    Step 2 — Check trailing stop: if BE already moved, trail SL behind swing points
    Step 3 — Check reversal (in profit): if reversal signals confirmed, exit early
    Step 4 — Check reversal (in drawdown): if aggressive drawdown with no reversal, exit early

    All actions are logged to Excel journal and sent via Telegram.
    """
    global trade_management_state

    # Fetch all currently open positions for this bot only (magic number 234000)
    positions = mt5.positions_get()
    if positions is None or len(positions) == 0:
        return   # no open trades — nothing to manage

    for pos in positions:
        if pos.magic != 234000:
            continue   # skip trades not placed by this bot

        ticket    = pos.ticket          # unique trade ID
        symbol    = pos.symbol
        direction = "BUY" if pos.type == mt5.ORDER_TYPE_BUY else "SELL"
        entry     = pos.price_open      # price trade was opened at
        current   = pos.price_current   # current market price
        sl        = pos.sl              # current stop loss
        tp        = pos.tp              # take profit
        volume    = pos.volume          # current lot size
        profit    = pos.profit          # floating P&L right now

        # Initialise state for this ticket if we haven't seen it before
        if ticket not in trade_management_state:
            trade_management_state[ticket] = {
                "partial_done": False,    # has 50% been closed?
                "be_done":      False,    # has SL been moved to break-even?
                "entry_time":   datetime.now(timezone.utc),
                "entry_price":  entry,
                "original_tp":  tp,
                "original_sl":  sl,
            }

        state = trade_management_state[ticket]

        # Calculate RR achieved so far (how far price has moved toward TP)
        if direction == "BUY":
            risk     = entry - sl if sl > 0 else 1        # distance from entry to SL
            progress = current - entry                     # how far price moved in our favour
        else:
            risk     = sl - entry if sl > 0 else 1
            progress = entry - current

        rr_now = progress / risk if risk > 0 else 0   # current RR ratio

        # ══════════════════════════════════════
        # STEP 1: PARTIAL PROFIT AT 1:1 RR
        # ══════════════════════════════════════
        if not state["partial_done"] and rr_now >= PARTIAL_CLOSE_RR:
            # Close 50% of the position at current market price
            partial_volume = round(volume * PARTIAL_CLOSE_PCT, 2)

            # Build the close request for the partial lot
            tick = mt5.symbol_info_tick(symbol)
            if tick:
                close_price = tick.bid if direction == "BUY" else tick.ask
                fill_mode   = get_filling_mode(symbol)

                partial_req = {
                    "action":       mt5.TRADE_ACTION_DEAL,
                    "symbol":       symbol,
                    "volume":       partial_volume,
                    "type":         mt5.ORDER_TYPE_SELL if direction == "BUY" else mt5.ORDER_TYPE_BUY,
                    "position":     ticket,          # link to open position
                    "price":        close_price,
                    "deviation":    20,
                    "magic":        234000,
                    "comment":      "V4 Partial 50%",
                    "type_time":    mt5.ORDER_TIME_GTC,
                    "type_filling": fill_mode,
                }

                result = mt5.order_send(partial_req)

                if result and result.retcode == mt5.TRADE_RETCODE_DONE:
                    state["partial_done"] = True    # mark partial as completed
                    print(f"✅ PARTIAL CLOSE {symbol} #{ticket} — {partial_volume} lots at {close_price}")

                    # Update journal with partial info
                    update_trade_in_journal(
                        symbol=symbol, timeframe="",
                        entry_price=state["entry_price"],
                        partial_taken="YES",
                        partial_price=close_price,
                    )

                    send_telegram(
                        f"💰 <b>PARTIAL PROFIT TAKEN — {symbol}</b>\n\n"
                        f"📊 Closed {partial_volume} lots at {close_price}\n"
                        f"📈 RR achieved: {rr_now:.2f}\n"
                        f"🔒 Runner continues — SL moving to break-even next\n"
                        f"💼 Locked profit: {profit * PARTIAL_CLOSE_PCT:.2f}"
                    )
                else:
                    err = result.comment if result else "No response"
                    print(f"❌ Partial close failed {symbol} #{ticket}: {err}")

        # ══════════════════════════════════════
        # STEP 2: MOVE SL TO BREAK-EVEN
        # After partial is done, move remaining SL to entry price
        # ══════════════════════════════════════
        if state["partial_done"] and not state["be_done"]:
            # Set new SL exactly at entry price — runner is now risk-free
            new_sl = entry

            be_req = {
                "action":   mt5.TRADE_ACTION_SLTP,   # modify SL/TP only
                "position": ticket,
                "sl":       new_sl,
                "tp":       tp,
            }

            result = mt5.order_send(be_req)

            if result and result.retcode == mt5.TRADE_RETCODE_DONE:
                state["be_done"] = True    # mark break-even as done
                print(f"🔒 BREAK-EVEN SET {symbol} #{ticket} — SL moved to {new_sl}")

                update_trade_in_journal(
                    symbol=symbol, timeframe="",
                    entry_price=state["entry_price"],
                    be_moved="YES",
                )

                send_telegram(
                    f"🔒 <b>BREAK-EVEN SET — {symbol}</b>\n\n"
                    f"🛑 SL moved to entry: {new_sl}\n"
                    f"🏃 Runner is now risk-free — targeting full TP\n"
                    f"📈 Current RR: {rr_now:.2f}"
                )
            else:
                err = result.comment if result else "No response"
                print(f"❌ BE move failed {symbol} #{ticket}: {err}")

        # ══════════════════════════════════════
        # STEP 3: TRAIL STOP BEHIND SWING POINTS
        # Only after BE is set — trail upward for BUY, downward for SELL
        # ══════════════════════════════════════
        elif state["be_done"] and rr_now > PARTIAL_CLOSE_RR:
            new_trail_sl = get_trailing_sl(symbol, direction, sl)

            if new_trail_sl != sl:
                # Trail has moved — update the SL on MT5
                trail_req = {
                    "action":   mt5.TRADE_ACTION_SLTP,
                    "position": ticket,
                    "sl":       new_trail_sl,
                    "tp":       tp,
                }
                result = mt5.order_send(trail_req)
                if result and result.retcode == mt5.TRADE_RETCODE_DONE:
                    print(f"📈 TRAIL SL MOVED {symbol} #{ticket} — {sl:.5f} → {new_trail_sl:.5f}")

        # ══════════════════════════════════════
        # STEP 4: DYNAMIC EXIT ON REVERSAL
        # If trade is in profit AND reversal is confirmed — exit early
        # If trade is in drawdown AND no reversal toward TP — exit early
        # ══════════════════════════════════════
        reversal_confirmed = get_reversal_signals(symbol, direction)

        if rr_now > 0.3 and reversal_confirmed:
            # Trade is in profit (>0.3 RR) and reversal is confirmed — lock in profit now
            tick = mt5.symbol_info_tick(symbol)
            if tick:
                exit_price = tick.bid if direction == "BUY" else tick.ask
                fill_mode  = get_filling_mode(symbol)

                exit_req = {
                    "action":       mt5.TRADE_ACTION_DEAL,
                    "symbol":       symbol,
                    "volume":       pos.volume,    # close whatever remains
                    "type":         mt5.ORDER_TYPE_SELL if direction == "BUY" else mt5.ORDER_TYPE_BUY,
                    "position":     ticket,
                    "price":        exit_price,
                    "deviation":    20,
                    "magic":        234000,
                    "comment":      "V4 Early Exit - Reversal",
                    "type_time":    mt5.ORDER_TIME_GTC,
                    "type_filling": fill_mode,
                }

                result = mt5.order_send(exit_req)

                if result and result.retcode == mt5.TRADE_RETCODE_DONE:
                    mins_open = (datetime.now(timezone.utc) - state["entry_time"]).seconds / 60
                    risk_val  = abs(entry - state["original_sl"])
                    rr_final  = (abs(exit_price - entry) / risk_val) if risk_val > 0 else 0

                    print(f"🚪 EARLY EXIT (Profit) {symbol} #{ticket} at {exit_price} RR:{rr_final:.2f}")

                    update_trade_in_journal(
                        symbol=symbol, timeframe="",
                        entry_price=state["entry_price"],
                        exit_reason="Early-Profit",
                        exit_price=exit_price,
                        actual_pnl=profit,
                        time_in_trade_mins=mins_open,
                        rr_achieved=rr_final,
                    )

                    send_telegram(
                        f"🚪 <b>EARLY EXIT — Reversal Detected — {symbol}</b>\n\n"
                        f"✅ Exited in profit before reversal hits\n"
                        f"📈 Exit price: {exit_price}\n"
                        f"💰 P&L: {profit:.2f}\n"
                        f"📊 RR achieved: {rr_final:.2f}\n"
                        f"⏱ Time in trade: {mins_open:.0f} mins"
                    )

                    # Clean up management state for this ticket
                    del trade_management_state[ticket]

        elif rr_now < -0.4 and reversal_confirmed is False:
            # Trade is in drawdown (>0.4R loss) and no reversal toward TP visible
            # Exit early to limit loss — better to lose 0.5R than full 1R
            tick = mt5.symbol_info_tick(symbol)
            if tick:
                exit_price = tick.bid if direction == "BUY" else tick.ask
                fill_mode  = get_filling_mode(symbol)

                exit_req = {
                    "action":       mt5.TRADE_ACTION_DEAL,
                    "symbol":       symbol,
                    "volume":       pos.volume,
                    "type":         mt5.ORDER_TYPE_SELL if direction == "BUY" else mt5.ORDER_TYPE_BUY,
                    "position":     ticket,
                    "price":        exit_price,
                    "deviation":    20,
                    "magic":        234000,
                    "comment":      "V4 Early Exit - SL Save",
                    "type_time":    mt5.ORDER_TIME_GTC,
                    "type_filling": fill_mode,
                }

                result = mt5.order_send(exit_req)

                if result and result.retcode == mt5.TRADE_RETCODE_DONE:
                    mins_open = (datetime.now(timezone.utc) - state["entry_time"]).seconds / 60
                    risk_val  = abs(entry - state["original_sl"])
                    rr_final  = (abs(exit_price - entry) / risk_val) * -1 if risk_val > 0 else 0

                    print(f"🛑 EARLY SL EXIT {symbol} #{ticket} at {exit_price} — saved from full SL")

                    update_trade_in_journal(
                        symbol=symbol, timeframe="",
                        entry_price=state["entry_price"],
                        exit_reason="Early-SL",
                        exit_price=exit_price,
                        actual_pnl=profit,
                        time_in_trade_mins=mins_open,
                        rr_achieved=rr_final,
                    )

                    send_telegram(
                        f"🛑 <b>EARLY SL EXIT — {symbol}</b>\n\n"
                        f"⚠️ No reversal detected — exiting before full stop loss\n"
                        f"📉 Exit price: {exit_price}\n"
                        f"💸 Loss saved vs full SL: {profit:.2f}\n"
                        f"⏱ Time in trade: {mins_open:.0f} mins"
                    )

                    del trade_management_state[ticket]


# ═══════════════════════════════════════════
# ALL FUNCTIONS BELOW ARE IDENTICAL TO V3
# Only the main loop and run() function were updated
# to call manage_open_trades() and check_daily_drawdown()
# ═══════════════════════════════════════════

def log_trade(
    symbol: str, timeframe: str, tier: str, confidence_score: int,
    entry_price: float, sl_price: float, tp_price: float,
    trigger_type: str, sl_source: str, session: str, status: str,
) -> None:
    """
    Wrapper: logs trade to Excel journal. Replaces old CSV log_trade.
    Called every time a signal fires before executing or alerting.
    """
    log_trade_excel(
        symbol=symbol, timeframe=timeframe, tier=tier,
        confidence_score=confidence_score, entry_price=entry_price,
        sl_price=sl_price, tp_price=tp_price,
        trigger_type=trigger_type, sl_source=sl_source,
        session=session, status=status,
    )


def send_telegram(message: str) -> None:
    """Send a message to the configured Telegram chat via bot API."""
    url     = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {"chat_id": TELEGRAM_CHAT_ID, "text": message, "parse_mode": "HTML"}
    try:
        requests.post(url, json=payload, timeout=10)
    except Exception as e:
        print(f"Telegram error: {e}")


def get_session() -> str:
    """Return current trading session name based on UTC hour."""
    h = datetime.now(timezone.utc).hour
    if   LONDON_START  <= h < LONDON_END:  return "London"
    elif OVERLAP_START <= h < OVERLAP_END: return "Overlap"
    elif NY_START      <= h < NY_END:      return "New York"
    elif POSTNY_START  <= h < POSTNY_END:  return "Post-NY"
    else:                                  return "Asian"


def is_active_session() -> bool:
    """Return True if currently in a tradeable session."""
    return get_session() != "Asian"


def is_bst() -> bool:
    """Detect whether UK is currently on British Summer Time (BST)."""
    now  = datetime.now(timezone.utc)
    year = now.year
    march_end = datetime(year, 4, 1, tzinfo=timezone.utc)
    bst_start = march_end - timedelta(days=(march_end.weekday() + 1) % 7)
    bst_start = bst_start.replace(hour=1, minute=0, second=0, microsecond=0)
    oct_end   = datetime(year, 11, 1, tzinfo=timezone.utc)
    bst_end   = oct_end - timedelta(days=(oct_end.weekday() + 1) % 7)
    bst_end   = bst_end.replace(hour=1, minute=0, second=0, microsecond=0)
    return bst_start <= now < bst_end


def get_dst_str() -> str:
    """Return 'BST' if summer time, else 'GMT'."""
    return "BST" if is_bst() else "GMT"


def get_pair_bias_hours(symbol: str) -> tuple:
    """Return display bias hours tuple for a symbol."""
    if symbol in THREE_AM_PAIRS:
        return (2, 6)
    return (1, 5)


def is_news_blackout() -> tuple:
    """Check if current time is within ±30 min of NFP, CPI, or FOMC."""
    now = datetime.now(timezone.utc)
    if now.weekday() == 4 and now.day <= 7:
        nfp_today = now.replace(hour=13, minute=30, second=0, microsecond=0)
        if abs(now - nfp_today) <= NEWS_WINDOW:
            return True, "NFP ±30min"
    for dt in CPI_DATES:
        if abs(now - dt) <= NEWS_WINDOW:
            return True, "CPI ±30min"
    for dt in FOMC_DATES:
        if abs(now - dt) <= NEWS_WINDOW:
            return True, "FOMC ±30min"
    return False, "Clear"


def connect_mt5() -> bool:
    """Initialise MT5 connection using .env credentials."""
    if not mt5.initialize(login=MT5_LOGIN, password=MT5_PASSWORD, server=MT5_SERVER):
        print("MT5 connection failed:", mt5.last_error())
        return False
    return True


def get_filling_mode(symbol: str) -> int:
    """Detect the correct order filling mode for a symbol."""
    info = mt5.symbol_info(symbol)
    if info is None:
        return mt5.ORDER_FILLING_IOC
    filling = info.filling_mode
    if filling & 2:
        return mt5.ORDER_FILLING_IOC
    elif filling & 1:
        return mt5.ORDER_FILLING_FOK
    else:
        return mt5.ORDER_FILLING_RETURN


def calc_lot_size(symbol: str, risk_amount_account: float, sl_distance_price: float) -> float:
    """Calculate lot size to risk exactly the specified amount."""
    info = mt5.symbol_info(symbol)
    if info is None or sl_distance_price <= 0:
        return 0.01
    tick_size  = info.trade_tick_size
    tick_value = info.trade_tick_value
    if tick_size <= 0 or tick_value <= 0:
        return 0.01
    ticks_in_sl   = sl_distance_price / tick_size
    value_per_lot = ticks_in_sl * tick_value
    if value_per_lot <= 0:
        return 0.01
    lot = risk_amount_account / value_per_lot
    lot = max(info.volume_min,
              min(info.volume_max,
                  round(lot / info.volume_step) * info.volume_step))
    return round(lot, 2)


def get_candles(timeframe: int, count: int = 500, symbol: str = None) -> pd.DataFrame:
    """Fetch OHLCV candle data from MT5 as a pandas DataFrame."""
    sym   = symbol or SYMBOL
    rates = mt5.copy_rates_from_pos(sym, timeframe, 0, count)
    if rates is None or len(rates) == 0:
        return None
    df = pd.DataFrame(rates)
    df['time'] = pd.to_datetime(df['time'], unit='s', utc=True)
    return df


def calc_ema(series: pd.Series, period: int) -> pd.Series:
    """Calculate EMA using pandas ewm — matches TradingView exactly."""
    return series.ewm(span=period, adjust=False).mean()


def calc_rsi(series: pd.Series, period: int = 14) -> pd.Series:
    """Calculate RSI using Wilder's smoothing."""
    delta    = series.diff()
    gain     = delta.where(delta > 0, 0)
    loss     = -delta.where(delta < 0, 0)
    avg_gain = gain.ewm(span=period).mean()
    avg_loss = loss.ewm(span=period).mean()
    rs       = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def calc_atr(df: pd.DataFrame, period: int = 14) -> pd.Series:
    """Calculate Average True Range."""
    hl = df['high'] - df['low']
    hc = abs(df['high'] - df['close'].shift())
    lc = abs(df['low']  - df['close'].shift())
    tr = pd.concat([hl, hc, lc], axis=1).max(axis=1)
    return tr.ewm(span=period).mean()


def calc_adx(df: pd.DataFrame, period: int = 14) -> pd.Series:
    """Calculate ADX — trend strength indicator."""
    plus_dm  = df['high'].diff()
    minus_dm = df['low'].diff().abs()
    plus_dm[plus_dm   < 0] = 0
    minus_dm[minus_dm < 0] = 0
    atr_val  = calc_atr(df, period)
    plus_di  = 100 * (plus_dm.ewm(span=period).mean()  / atr_val)
    minus_di = 100 * (minus_dm.ewm(span=period).mean() / atr_val)
    dx       = (abs(plus_di - minus_di) / (plus_di + minus_di)) * 100
    return dx.ewm(span=period).mean()


def calc_macd(series: pd.Series) -> tuple:
    """Calculate MACD line and signal line."""
    macd   = series.ewm(span=12).mean() - series.ewm(span=26).mean()
    signal = macd.ewm(span=9).mean()
    return macd, signal


def is_strong_body(
    candle_open: float, candle_close: float,
    candle_high: float, candle_low: float,
    threshold: float = 0.60,
) -> tuple:
    """Check if candle body covers at least 60% of total range."""
    c_range = candle_high - candle_low
    if c_range <= 0:
        return False, False, False
    c_body      = abs(candle_close - candle_open)
    strong      = (c_body / c_range) >= threshold
    bull_strong = strong and candle_close > candle_open
    bear_strong = strong and candle_close < candle_open
    return strong, bull_strong, bear_strong


def get_weekly_bias(df_4h: pd.DataFrame, symbol: str = "XAUUSD") -> tuple:
    """Calculate weekly bias from Monday's 00:00 and 04:00 UTC 4H candles."""
    df = df_4h.copy()
    df['dow']  = df['time'].dt.dayofweek
    df['hour'] = df['time'].dt.hour
    if len(df) == 0:
        return None, None, False, False, False, False, False
    latest_week = df['time'].dt.isocalendar().week.iloc[-1]
    latest_year = df['time'].dt.isocalendar().year.iloc[-1]
    this_week   = df[
        (df['time'].dt.isocalendar().week == latest_week) &
        (df['time'].dt.isocalendar().year == latest_year)
    ]
    if len(this_week) == 0:
        return None, None, False, False, False, False, False
    c1_rows = this_week[(this_week['dow'] == 0) & (this_week['hour'] == BIAS_CANDLE1_HOUR)]
    c2_rows = this_week[(this_week['dow'] == 0) & (this_week['hour'] == BIAS_CANDLE2_HOUR)]
    if len(c1_rows) == 0:
        return None, None, False, False, False, False, False
    c1 = c1_rows.iloc[-1]
    wb_high = wb_low = None
    wk_partial = wk_full = False
    if len(c2_rows) == 0:
        wb_high = c1['high']
        wb_low  = c1['low']
        wk_partial = True
    else:
        c2      = c2_rows.iloc[-1]
        wb_high = max(c1['high'], c2['high'])
        wb_low  = min(c1['low'],  c2['low'])
        wk_full = True
    current_close  = df['close'].iloc[-1]
    price_in_range = wb_low < current_close < wb_high
    wk_bull = wk_bear = False
    if wk_full and len(c2_rows) > 0:
        c2_time         = c2_rows.iloc[-1]['time']
        confirm_candles = this_week[this_week['time'] > c2_time]
        for _, row in confirm_candles.iterrows():
            _, bull_strong, bear_strong = is_strong_body(row['open'], row['close'], row['high'], row['low'])
            if not wk_bull and row['close'] > wb_high and bull_strong:
                wk_bull = True
            if not wk_bear and row['close'] < wb_low and bear_strong:
                wk_bear = True
            if wk_bull or wk_bear:
                break
    return wb_high, wb_low, wk_bull, wk_bear, price_in_range, wk_partial, wk_full


def get_daily_bias(df_4h: pd.DataFrame, symbol: str = "XAUUSD") -> tuple:
    """Calculate daily bias from 00:00 UTC 4H candle. No fallback — matches Pine Script."""
    df = df_4h.copy()
    df['date'] = df['time'].dt.date
    df['hour'] = df['time'].dt.hour
    today         = df['date'].iloc[-1]
    today_candles = df[df['date'] == today]
    c1_rows = today_candles[today_candles['hour'] == BIAS_CANDLE1_HOUR]
    if len(c1_rows) == 0:
        return None, None, False, False
    c1      = c1_rows.iloc[-1]
    db_high = c1['high']
    db_low  = c1['low']
    confirm_candles = today_candles[today_candles['time'] > c1['time']]
    day_bull = day_bear = False
    for _, row in confirm_candles.iterrows():
        _, bull_strong, bear_strong = is_strong_body(row['open'], row['close'], row['high'], row['low'])
        if not day_bull and row['close'] > db_high and bull_strong:
            day_bull = True
        if not day_bear and row['close'] < db_low and bear_strong:
            day_bear = True
        if day_bull or day_bear:
            break
    return db_high, db_low, day_bull, day_bear


def get_htf_ema(df_1h: pd.DataFrame, df_4h: pd.DataFrame) -> tuple:
    """Check 1H and 4H closes vs EMA200 for HTF trend confirmation."""
    ema200_1h = calc_ema(df_1h['close'], 200).iloc[-1]
    ema200_4h = calc_ema(df_4h['close'], 200).iloc[-1]
    c1h = df_1h['close'].iloc[-1]
    c4h = df_4h['close'].iloc[-1]
    htf_bull = (c1h > ema200_1h) or (c4h > ema200_4h)
    htf_bear = (c1h < ema200_1h) or (c4h < ema200_4h)
    return htf_bull, htf_bear


def get_prev_day_hl(df: pd.DataFrame) -> tuple:
    """Get previous day high and low."""
    df = df.copy()
    df['date'] = df['time'].dt.date
    dates = sorted(df['date'].unique())
    if len(dates) < 2:
        return None, None
    prev = df[df['date'] == dates[-2]]
    return prev['high'].max(), prev['low'].min()


def get_prev_week_hl(df: pd.DataFrame) -> tuple:
    """Get previous week high and low."""
    df = df.copy()
    df['week'] = df['time'].dt.isocalendar().week
    df['year'] = df['time'].dt.isocalendar().year
    weeks = df[['year','week']].drop_duplicates().values.tolist()
    if len(weeks) < 2:
        return None, None
    py, pw = weeks[-2]
    prev = df[(df['year'] == py) & (df['week'] == pw)]
    return prev['high'].max(), prev['low'].min()


def get_prev_month_hl(df: pd.DataFrame) -> tuple:
    """Get previous month high and low."""
    df = df.copy()
    df['month'] = df['time'].dt.month
    df['year']  = df['time'].dt.year
    months = df[['year','month']].drop_duplicates().values.tolist()
    if len(months) < 2:
        return None, None
    py, pm = months[-2]
    prev = df[(df['year'] == py) & (df['month'] == pm)]
    return prev['high'].max(), prev['low'].min()


def get_bos_reset(timeframe: int) -> int:
    """Return timeframe-appropriate BOS lookback window."""
    if timeframe == mt5.TIMEFRAME_M5:
        return 50
    elif timeframe in (mt5.TIMEFRAME_M15, mt5.TIMEFRAME_M30):
        return 30
    else:
        return 20


def check_bos(df: pd.DataFrame, atr_series: pd.Series, timeframe: int) -> tuple:
    """Detect recent Break of Structure with timeframe-aware lookback."""
    bos_reset   = get_bos_reset(timeframe)
    close       = df['close']
    open_       = df['open']
    ema200      = calc_ema(close, 200)
    sw_high     = df['high'].rolling(11, center=True).max()
    sw_low      = df['low'].rolling(11,  center=True).min()
    strong_body = abs(close - open_) > atr_series * 0.5
    above_e200  = close > ema200
    below_e200  = close < ema200
    bull_bos = (close > sw_high.shift(1)) & strong_body & above_e200 & (close.shift(1) <= sw_high.shift(1))
    bear_bos = (close < sw_low.shift(1))  & strong_body & below_e200 & (close.shift(1) >= sw_low.shift(1))
    return bull_bos.iloc[-bos_reset:].any(), bear_bos.iloc[-bos_reset:].any()


def check_sweeps(df: pd.DataFrame, atr_series: pd.Series, d1h: float, d1l: float) -> tuple:
    """Detect liquidity sweeps above/below previous day high/low."""
    if d1h is None or d1l is None:
        return False, False, False, False, None, None
    last = df.iloc[-2]
    atr  = atr_series.iloc[-2]
    bull_sweep  = (last['low'] < d1l)  and (last['close'] > d1l)  and ((d1l - last['low'])   >= atr * 0.3)
    bear_sweep  = (last['high'] > d1h) and (last['close'] < d1h)  and ((last['high'] - d1h)  >= atr * 0.3)
    bull_sw_rej = bull_sweep and ((last['close'] - last['low'])   > atr * 0.7)
    bear_sw_rej = bear_sweep and ((last['high']  - last['close']) > atr * 0.7)
    return (bull_sweep, bear_sweep, bull_sw_rej, bear_sw_rej,
            last['low']  if bull_sweep else None,
            last['high'] if bear_sweep else None)


def check_internal_sweep(df: pd.DataFrame, above_e200: bool, below_e200: bool) -> tuple:
    """Detect short-term internal liquidity grabs."""
    low3  = df['low'].iloc[-4:-1].min()
    high3 = df['high'].iloc[-4:-1].max()
    last  = df.iloc[-2]
    int_bull = (last['low'] < low3)   and (last['close'] > low3)  and above_e200
    int_bear = (last['high'] > high3) and (last['close'] < high3) and below_e200
    return int_bull, int_bear


def check_eql_swept(df: pd.DataFrame, atr: float) -> tuple:
    """Detect equal highs/lows sweep."""
    ph  = df['high'].iloc[-12]
    pph = df['high'].iloc[-23]
    pl  = df['low'].iloc[-12]
    ppl = df['low'].iloc[-23]
    last    = df.iloc[-2]
    is_eqh  = (abs(ph - pph) <= atr * 0.1) if not (pd.isna(ph) or pd.isna(pph)) else False
    is_eql  = (abs(pl - ppl) <= atr * 0.1) if not (pd.isna(pl) or pd.isna(ppl)) else False
    eqh_swept = is_eqh and (last['high'] > ph) and (last['close'] < ph)
    eql_swept = is_eql and (last['low']  < pl) and (last['close'] > pl)
    return eqh_swept, eql_swept


def check_fvg(
    df, atr_series, above_e200, below_e200,
    wk_bull, wk_bear, day_bull, day_bear,
    bias_med_bull, bias_med_bear, bias_sca_bull, bias_sca_bear,
) -> tuple:
    """Detect Fair Value Gap entry zones."""
    if len(df) < 6:
        return False, False, None, None
    c0  = df.iloc[-1]
    c1  = df.iloc[-2]
    c2  = df.iloc[-3]
    c3  = df.iloc[-4]
    atr   = atr_series.iloc[-3]
    body  = abs(c2['close'] - c2['open'])
    str_cdl = body > atr * 1.0
    b_fvg_sz  = c1['low']  - c3['high']
    br_fvg_sz = c3['low']  - c1['high']
    fvg_min = 0.5
    b_fvg   = (b_fvg_sz  > 0) and str_cdl and (c2['close'] > c2['open']) and above_e200 and (b_fvg_sz  >= atr * fvg_min)
    br_fvg  = (br_fvg_sz > 0) and str_cdl and (c2['close'] < c2['open']) and below_e200 and (br_fvg_sz >= atr * fvg_min)
    disp_cdl = abs(c0['close'] - c0['open']) > atr_series.iloc[-1] * 1.5
    bull_in_fvg = b_fvg  and (c0['low'] <= c1['low'])   and (c0['close'] >= c3['high'])
    bear_in_fvg = br_fvg and (c0['high'] >= c1['high'])  and (c0['close'] <= c3['low'])
    bull_fvg_ez = bull_in_fvg and disp_cdl and above_e200 and (wk_bull or bias_med_bull or bias_sca_bull)
    bear_fvg_ez = bear_in_fvg and disp_cdl and below_e200 and (wk_bear or bias_med_bear or bias_sca_bear)
    bull_fvg_bot = c3['high'] if b_fvg else None
    bear_fvg_top = c3['low']  if br_fvg else None
    return bull_fvg_ez, bear_fvg_ez, bull_fvg_bot, bear_fvg_top


def check_ob_stateful(
    df, atr_series, above_e200, below_e200, ob_state, key,
) -> tuple:
    """Detect and track Order Blocks with persistent state. OB at iloc[-5]."""
    if len(df) < 8:
        return False, False, None, None
    close  = df['close']
    open_  = df['open']
    atr    = atr_series.iloc[-2]
    n      = len(df)
    bull_imp = (close.iloc[-2] > close.iloc[-3] and close.iloc[-3] > close.iloc[-4] and close.iloc[-4] > close.iloc[-5])
    bear_imp = (close.iloc[-2] < close.iloc[-3] and close.iloc[-3] < close.iloc[-4] and close.iloc[-4] < close.iloc[-5])
    if bull_imp and above_e200 and len(df) >= 6:
        ob_candle = df.iloc[-5]
        if (ob_candle['close'] < ob_candle['open'] and abs(ob_candle['close'] - ob_candle['open']) > atr * 0.5):
            existing = ob_state[key].get('bull_ob')
            if not existing or not existing.get('active', False):
                ob_state[key]['bull_ob'] = {'top': ob_candle['high'], 'bot': ob_candle['low'], 'active': True, 'touched': False, 'bar_formed': n, 'formed_at': str(ob_candle['time'])}
    if bear_imp and below_e200 and len(df) >= 6:
        ob_candle = df.iloc[-5]
        if (ob_candle['close'] > ob_candle['open'] and abs(ob_candle['close'] - ob_candle['open']) > atr * 0.5):
            existing = ob_state[key].get('bear_ob')
            if not existing or not existing.get('active', False):
                ob_state[key]['bear_ob'] = {'top': ob_candle['high'], 'bot': ob_candle['low'], 'active': True, 'touched': False, 'bar_formed': n, 'formed_at': str(ob_candle['time'])}
    bull_ob_rsp = bear_ob_rsp = False
    bull_ob_bot = bear_ob_top = None
    current = df.iloc[-2]
    bob = ob_state[key].get('bull_ob')
    if bob and bob.get('active') and not bob.get('touched'):
        if current['low'] <= bob['top'] and current['close'] >= bob['bot']:
            bull_ob_rsp = True
            bull_ob_bot = bob['bot']
            ob_state[key]['bull_ob']['touched'] = True
        if current['close'] < bob['bot']:
            ob_state[key]['bull_ob']['active'] = False
    bearob = ob_state[key].get('bear_ob')
    if bearob and bearob.get('active') and not bearob.get('touched'):
        if current['high'] >= bearob['bot'] and current['close'] <= bearob['top']:
            bear_ob_rsp = True
            bear_ob_top = bearob['top']
            ob_state[key]['bear_ob']['touched'] = True
        if current['close'] > bearob['top']:
            ob_state[key]['bear_ob']['active'] = False
    if bob and (n - bob.get('bar_formed', 0)) > 50:
        ob_state[key]['bull_ob']['active'] = False
    if bearob and (n - bearob.get('bar_formed', 0)) > 50:
        ob_state[key]['bear_ob']['active'] = False
    return bull_ob_rsp, bear_ob_rsp, bull_ob_bot, bear_ob_top


def check_rsi_div(df: pd.DataFrame, rsi_series: pd.Series) -> tuple:
    """Detect RSI divergence patterns."""
    lkb  = 7
    low  = df['low']
    high = df['high']
    p_ll = low.iloc[-lkb:].min()   < low.iloc[-2*lkb:-lkb].min()
    p_hl = low.iloc[-lkb:].min()   > low.iloc[-2*lkb:-lkb].min()
    p_hh = high.iloc[-lkb:].max()  > high.iloc[-2*lkb:-lkb].max()
    p_lh = high.iloc[-lkb:].max()  < high.iloc[-2*lkb:-lkb].max()
    r_hl = rsi_series.iloc[-lkb:].min() > rsi_series.iloc[-2*lkb:-lkb].min()
    r_ll = rsi_series.iloc[-lkb:].min() < rsi_series.iloc[-2*lkb:-lkb].min()
    r_lh = rsi_series.iloc[-lkb:].max() < rsi_series.iloc[-2*lkb:-lkb].max()
    r_hh = rsi_series.iloc[-lkb:].max() > rsi_series.iloc[-2*lkb:-lkb].max()
    return (p_ll and r_hl), (p_hl and r_ll), (p_hh and r_lh), (p_lh and r_hh)


def calc_structure_sl_tp(
    close, atr, t1_bull, t2_bull, t1_bear, t2_bear,
    bull_ob_bot, bear_ob_top, bull_fvg_bot, bear_fvg_top,
    df, wb_high, wb_low, bull_tier, bear_tier, rr=3.0,
) -> tuple:
    """Calculate SL and TP from market structure. Matches Pine Script exactly."""
    session = get_session()
    if session in ("London", "New York", "Overlap"):
        sess_buf = atr * 0.2
    elif session == "Post-NY":
        sess_buf = atr * 0.3
    else:
        sess_buf = atr * 0.5
    local_swing_low  = df['low'].iloc[-17:-2].min()
    local_swing_high = df['high'].iloc[-17:-2].max()
    long_sl_atr  = close - atr * 1.5
    short_sl_atr = close + atr * 1.5
    if t2_bull and bull_ob_bot is not None:
        long_sl = bull_ob_bot - sess_buf; sl_src_bull = "OB"
    elif t1_bull and bull_fvg_bot is not None:
        long_sl = bull_fvg_bot - sess_buf; sl_src_bull = "FVG"
    elif (local_swing_low - sess_buf) > (close - atr * 3):
        long_sl = local_swing_low - sess_buf; sl_src_bull = "Swing"
    else:
        long_sl = long_sl_atr; sl_src_bull = "ATR"
    long_risk = max(close - long_sl, 0.0001)
    if t2_bear and bear_ob_top is not None:
        short_sl = bear_ob_top + sess_buf; sl_src_bear = "OB"
    elif t1_bear and bear_fvg_top is not None:
        short_sl = bear_fvg_top + sess_buf; sl_src_bear = "FVG"
    elif (local_swing_high + sess_buf) < (close + atr * 3):
        short_sl = local_swing_high + sess_buf; sl_src_bear = "Swing"
    else:
        short_sl = short_sl_atr; sl_src_bear = "ATR"
    short_risk   = max(short_sl - close, 0.0001)
    long_tp_base  = close + long_risk  * rr
    short_tp_base = close - short_risk * rr
    if bull_tier == "STRONG":
        long_tp = long_tp_base
    elif bull_tier == "MEDIUM":
        long_tp = wb_high if (wb_high is not None and wb_high >= long_tp_base) else long_tp_base
    elif bull_tier == "SCALP":
        long_tp_sca = wb_high if (wb_high is not None and wb_high >= long_tp_base) else long_tp_base
        long_rr_sca = (long_tp_sca - close) / long_risk if long_risk > 0 else 0
        long_tp     = long_tp_sca if long_rr_sca >= SCALP_MIN_RR else long_tp_base
    else:
        long_tp = long_tp_base
    if bear_tier == "STRONG":
        short_tp = short_tp_base
    elif bear_tier == "MEDIUM":
        short_tp = wb_low if (wb_low is not None and wb_low <= short_tp_base) else short_tp_base
    elif bear_tier == "SCALP":
        short_tp_sca = wb_low if (wb_low is not None and wb_low <= short_tp_base) else short_tp_base
        short_rr_sca = (close - short_tp_sca) / short_risk if short_risk > 0 else 0
        short_tp     = short_tp_sca if short_rr_sca >= SCALP_MIN_RR else short_tp_base
    else:
        short_tp = short_tp_base
    return (round(long_sl, 5), long_risk, round(long_tp, 5), sl_src_bull,
            round(short_sl, 5), short_risk, round(short_tp, 5), sl_src_bear)


def check_scalp_rr_ok(close: float, tp: float, sl: float, tier: str) -> bool:
    """Block scalp signals where RR < 1.5."""
    if tier != "SCALP":
        return True
    risk = abs(close - sl)
    if risk <= 0:
        return False
    return (abs(tp - close) / risk) >= SCALP_MIN_RR


def calc_confidence_score(
    wk_bull, wk_bear, day_bull, day_bear, htf_bull, htf_bear,
    active_session, ext_bear, s2_long, s2_short,
    any_trig_bull, any_trig_bear, bon_l, bon_s,
    min_str2, min_bonus, is_bull_signal,
) -> int:
    """Calculate 0-10 confidence score for a signal direction."""
    score = 0
    if wk_bull or wk_bear:          score += 1
    if day_bull or day_bear:        score += 1
    bias_strong = (wk_bull and day_bull) or (wk_bear and day_bear)
    bias_medium = (wk_bull and day_bear) or (wk_bear and day_bull)
    if bias_strong:   score += 2
    elif bias_medium: score += 1
    if (is_bull_signal and htf_bull) or (not is_bull_signal and htf_bear): score += 1
    if active_session:  score += 1
    if not ext_bear:    score += 1
    s2 = s2_long if is_bull_signal else s2_short
    if s2 >= min_str2:  score += 1
    trig_active = any_trig_bull if is_bull_signal else any_trig_bear
    if trig_active:     score += 1
    bon = bon_l if is_bull_signal else bon_s
    if bon >= min_bonus: score += 1
    return score


def get_session_threshold(session: str) -> int:
    """Return confidence threshold for current session."""
    return CONFIDENCE_THRESHOLD_POSTNY if session == "Post-NY" else CONFIDENCE_THRESHOLD_PRIME


def place_trade(
    signal: str, sl_price: float, tp_price: float,
    lot_size: float, symbol: str = None, tf_name: str = "30M",
) -> None:
    """Submit a market order to MT5 with SL and TP."""
    sym  = symbol or SYMBOL
    tick = mt5.symbol_info_tick(sym)
    if tick is None:
        send_telegram(f"❌ Failed to get price for {sym}!")
        return
    price      = tick.ask if signal == "BUY" else tick.bid
    order_type = mt5.ORDER_TYPE_BUY if signal == "BUY" else mt5.ORDER_TYPE_SELL
    filling    = get_filling_mode(sym)
    req = {
        "action":       mt5.TRADE_ACTION_DEAL,
        "symbol":       sym,
        "volume":       lot_size,
        "type":         order_type,
        "price":        price,
        "sl":           round(sl_price, 5),
        "tp":           round(tp_price, 5),
        "deviation":    20,
        "magic":        234000,
        "comment":      f"Mannys V4 {tf_name}",
        "type_time":    mt5.ORDER_TIME_GTC,
        "type_filling": filling,
    }
    result  = mt5.order_send(req)
    account = mt5.account_info()
    if result is None or result.retcode != mt5.TRADE_RETCODE_DONE:
        err = result.comment if result else "No response from MT5"
        send_telegram(f"❌ <b>TRADE FAILED</b>\n{sym} {tf_name} {signal}\nError: {err}")
        print(f"❌ Trade failed: {err}")
    else:
        dst_str = get_dst_str()
        send_telegram(
            f"🥇 <b>TRADE PLACED — Manny's V4</b>\n\n"
            f"📊 {signal} | {sym} | {tf_name}\n"
            f"🏦 {get_session()} | {dst_str}\n"
            f"📈 Entry: {result.price}\n"
            f"🛑 SL: {round(sl_price,5)}\n"
            f"🎯 TP: {round(tp_price,5)}\n"
            f"📦 Lots: {lot_size}\n"
            f"💼 Balance: {account.balance} GBP\n"
            f"⏰ {datetime.now(timezone.utc).strftime('%H:%M:%S')} UTC"
        )
        print(f"✅ {sym} {tf_name} {signal} | Entry:{result.price} SL:{round(sl_price,5)} TP:{round(tp_price,5)}")


# ═══════════════════════════════════════════
# MAIN SIGNAL CHECK — identical to V3
# ═══════════════════════════════════════════
def check_signal(
    last_signal_bar, last_bull_bar, last_bear_bar,
    symbol: str = None, timeframe: int = None, ob_state: dict = None,
) -> tuple:
    """
    Core signal evaluation — unchanged from V3.
    All Tier 1 trade management happens in manage_open_trades() separately.
    """
    sym     = symbol or SYMBOL
    tf      = timeframe or mt5.TIMEFRAME_M30
    tf_name = TIMEFRAME_NAMES.get(tf, "30M")
    key     = f"{sym}_{tf_name}"
    if ob_state is None:
        ob_state = {}
    if key not in ob_state:
        ob_state[key] = {'bull_ob': None, 'bear_ob': None}

    news_black, news_reason = is_news_blackout()
    if news_black:
        print(f"  🚫 {sym} {tf_name} — NEWS BLACKOUT: {news_reason}")
        return last_signal_bar, last_bull_bar, last_bear_bar

    df_main = get_candles(tf, 500, sym)
    df_4h   = get_candles(TIMEFRAME_4H, 500, sym)
    df_1h   = get_candles(TIMEFRAME_1H, 500, sym)

    if df_main is None or df_4h is None or df_1h is None:
        print(f"  ⚠ No data: {sym} {tf_name}")
        return last_signal_bar, last_bull_bar, last_bear_bar
    if len(df_main) < 210 or len(df_4h) < 18:
        print(f"  ⚠ Not enough bars: {sym} {tf_name}")
        return last_signal_bar, last_bull_bar, last_bear_bar

    df_main['ema200'] = calc_ema(df_main['close'], 200)
    df_main['ema50']  = calc_ema(df_main['close'], 50)
    df_main['rsi']    = calc_rsi(df_main['close'], 14)
    df_main['atr']    = calc_atr(df_main, ATR_PERIOD)
    df_main['adx']    = calc_adx(df_main, ADX_PERIOD)
    df_main['macd'], df_main['macd_sig'] = calc_macd(df_main['close'])

    last        = df_main.iloc[-2]
    current_bar = last['time']
    bar_index   = len(df_main) - 2

    tf_cooldown = {mt5.TIMEFRAME_M5: 12, mt5.TIMEFRAME_M15: 10, mt5.TIMEFRAME_M30: 8, mt5.TIMEFRAME_H1: 5, mt5.TIMEFRAME_H4: 3}.get(tf, 8)
    bars_since_bull = (bar_index - last_bull_bar) if last_bull_bar is not None else 999
    bars_since_bear = (bar_index - last_bear_bar) if last_bear_bar is not None else 999
    bull_cooled = bars_since_bull > tf_cooldown
    bear_cooled = bars_since_bear > tf_cooldown

    close   = last['close']
    ema200  = last['ema200']
    ema50   = last['ema50']
    rsi     = last['rsi']
    atr     = last['atr']
    adx_val = last['adx']
    macd_l  = last['macd']
    macd_s  = last['macd_sig']

    above_e200 = close > ema200
    below_e200 = close < ema200
    pct_below  = (ema200 - close) / ema200 * 100 if ema200 > 0 else 0
    ext_bear   = pct_below > 20
    trending   = adx_val >= ADX_THRESH
    macd_bull  = macd_l > macd_s
    macd_bear  = macd_l < macd_s
    rsi_bull_x = rsi < 40
    rsi_bear_x = rsi > 60

    wb_high, wb_low, wk_bull, wk_bear, price_in_weekly_range, wk_partial, wk_full = get_weekly_bias(df_4h, sym)
    db_high, db_low, day_bull, day_bear = get_daily_bias(df_4h, sym)
    htf_bull, htf_bear = get_htf_ema(df_1h, df_4h)

    bias_med_bull = wk_bear and day_bull
    bias_med_bear = wk_bull and day_bear
    bias_sca_bull = price_in_weekly_range and day_bull
    bias_sca_bear = price_in_weekly_range and day_bear

    macro_trend_bull = wk_bull and day_bull
    macro_trend_bear = wk_bear and day_bear

    g_str_bull = wk_bull and day_bull and htf_bull and (not ext_bear) and macro_trend_bull
    g_str_bear = wk_bear and day_bear and htf_bear and (not ext_bear) and macro_trend_bear
    g_med_bull = wk_bear and day_bull and htf_bull and (not ext_bear)
    g_med_bear = wk_bull and day_bear and htf_bear and (not ext_bear)
    g_sca_bull = price_in_weekly_range and day_bull and htf_bull and (not ext_bear)
    g_sca_bear = price_in_weekly_range and day_bear and htf_bear and (not ext_bear)

    d1h, d1l      = get_prev_day_hl(df_main)
    w_high, w_low = get_prev_week_hl(df_main)
    m_high, m_low = get_prev_month_hl(df_main)

    w_mid   = (w_low + (w_high - w_low) / 2) if w_high and w_low else None
    in_disc = (close < w_mid) if w_mid else False
    in_prem = (close > w_mid) if w_mid else False
    m_bull  = (close > m_low  and above_e200) if m_low  else False
    m_bear  = (close < m_high and below_e200) if m_high else False

    rec_bull_bos, rec_bear_bos = check_bos(df_main, df_main['atr'], tf)
    (bull_sweep, bear_sweep, bull_sw_rej, bear_sw_rej,
     last_bull_sw_low, last_bear_sw_high) = check_sweeps(df_main, df_main['atr'], d1h, d1l)
    int_bull, int_bear = check_internal_sweep(df_main, above_e200, below_e200)
    eqh_swept, eql_swept = check_eql_swept(df_main, atr)

    s2_long  = sum([above_e200, rec_bull_bos, bull_sweep, in_disc, m_bull])
    s2_short = sum([below_e200, rec_bear_bos, bear_sweep, in_prem, m_bear])

    bull_fvg_ez, bear_fvg_ez, bull_fvg_bot, bear_fvg_top = check_fvg(
        df_main, df_main['atr'], above_e200, below_e200,
        wk_bull, wk_bear, day_bull, day_bear,
        bias_med_bull, bias_med_bear, bias_sca_bull, bias_sca_bear
    )

    bull_ob_rsp, bear_ob_rsp, bull_ob_bot, bear_ob_top = check_ob_stateful(
        df_main, df_main['atr'], above_e200, below_e200, ob_state, key
    )

    s_bull_div, h_bull_div, s_bear_div, h_bear_div = check_rsi_div(df_main, df_main['rsi'])

    bull_wick = last['low']  - min(last['open'], last['close'])
    bear_wick = max(last['open'], last['close']) - last['high']
    pin_body  = abs(last['close'] - last['open'])

    t1_bull = bull_fvg_ez
    t1_bear = bear_fvg_ez
    t2_bull = bull_ob_rsp
    t2_bear = bear_ob_rsp
    t3_bull = bull_sw_rej
    t3_bear = bear_sw_rej
    t4_bull = ((bull_wick >= 2.5 * pin_body) and (pin_body >= atr * 0.1) and above_e200 and (wk_bull or bias_med_bull or bias_sca_bull) and day_bull)
    t4_bear = ((bear_wick >= 2.5 * pin_body) and (pin_body >= atr * 0.1) and below_e200 and (wk_bear or bias_med_bear or bias_sca_bear) and day_bear)
    t5_bull = ((d1h is not None) and (close > d1h) and (df_main['close'].iloc[-3] <= d1h) and above_e200 and (wk_bull or bias_med_bull or bias_sca_bull) and day_bull)
    t5_bear = ((d1l is not None) and (close < d1l) and (df_main['close'].iloc[-3] >= d1l) and below_e200 and (wk_bear or bias_med_bear or bias_sca_bear) and day_bear)
    lh1 = (df_main['high'].iloc[-2] < df_main['high'].iloc[-3] and df_main['high'].iloc[-3] < df_main['high'].iloc[-4] and df_main['high'].iloc[-4] < df_main['high'].iloc[-5])
    hl1 = (df_main['low'].iloc[-2]  > df_main['low'].iloc[-3]  and df_main['low'].iloc[-3]  > df_main['low'].iloc[-4]  and df_main['low'].iloc[-4]  > df_main['low'].iloc[-5])
    t7_bull = (lh1 and (last['high'] > df_main['high'].iloc[-3]) and above_e200 and (wk_bull or bias_med_bull or bias_sca_bull) and day_bull)
    t7_bear = (hl1 and (last['low']  < df_main['low'].iloc[-3])  and below_e200 and (wk_bear or bias_med_bear or bias_sca_bear) and day_bear)

    any_trig_bull = t1_bull or t2_bull or t3_bull or t4_bull or t5_bull or t7_bull
    any_trig_bear = t1_bear or t2_bear or t3_bear or t4_bear or t5_bear or t7_bear

    trig_name_bull = ("FVG" if t1_bull else "OB" if t2_bull else "Sweep" if t3_bull else "Pin" if t4_bull else "PDH" if t5_bull else "CHoCH" if t7_bull else "None")
    trig_name_bear = ("FVG" if t1_bear else "OB" if t2_bear else "Sweep" if t3_bear else "Pin" if t4_bear else "PDL" if t5_bear else "CHoCH" if t7_bear else "None")

    bon_l  = 0
    bon_l += 2 if s_bull_div else (1 if h_bull_div else 0)
    bon_l += 1 if macd_bull  else 0
    bon_l += 1 if rsi_bull_x else 0
    bon_l += 1 if trending   else 0
    bon_l += 1 if int_bull   else 0
    bon_l += 1 if eql_swept  else 0

    bon_s  = 0
    bon_s += 2 if s_bear_div else (1 if h_bear_div else 0)
    bon_s += 1 if macd_bear  else 0
    bon_s += 1 if rsi_bear_x else 0
    bon_s += 1 if trending   else 0
    bon_s += 1 if int_bear   else 0
    bon_s += 1 if eqh_swept  else 0

    min_bonus = 2
    min_str2  = 2

    raw_bull_str = (g_str_bull and bull_cooled and (s2_long  >= min_str2) and any_trig_bull and (bon_l >= min_bonus))
    raw_bear_str = (g_str_bear and bear_cooled and (s2_short >= min_str2) and any_trig_bear and (bon_s >= min_bonus))
    raw_bull_med = (g_med_bull and bull_cooled and (s2_long  >= min_str2) and any_trig_bull and (bon_l >= max(min_bonus-1,0)))
    raw_bear_med = (g_med_bear and bear_cooled and (s2_short >= min_str2) and any_trig_bear and (bon_s >= max(min_bonus-1,0)))
    raw_bull_sca = (g_sca_bull and bull_cooled and (s2_long  >= max(min_str2-1,1)) and any_trig_bull and (bon_l >= max(min_bonus-1,0)))
    raw_bear_sca = (g_sca_bear and bear_cooled and (s2_short >= max(min_str2-1,1)) and any_trig_bear and (bon_s >= max(min_bonus-1,0)))

    raw_bull = (raw_bull_str or (raw_bull_med and not raw_bull_str) or (raw_bull_sca and not raw_bull_str and not raw_bull_med))
    raw_bear = (raw_bear_str or (raw_bear_med and not raw_bear_str) or (raw_bear_sca and not raw_bear_str and not raw_bear_med))

    bull_tier = "STRONG" if raw_bull_str else ("MEDIUM" if raw_bull_med else ("SCALP" if raw_bull_sca else ""))
    bear_tier = "STRONG" if raw_bear_str else ("MEDIUM" if raw_bear_med else ("SCALP" if raw_bear_sca else ""))

    (long_sl, long_risk, long_tp, sl_src_bull,
     short_sl, short_risk, short_tp, sl_src_bear) = calc_structure_sl_tp(
        close, atr, t1_bull, t2_bull, t1_bear, t2_bear,
        bull_ob_bot, bear_ob_top, bull_fvg_bot, bear_fvg_top,
        df_main, wb_high, wb_low, bull_tier, bear_tier, rr=RR
    )

    bull_sig_final = raw_bull and check_scalp_rr_ok(close, long_tp, long_sl, bull_tier)
    bear_sig_final = raw_bear and check_scalp_rr_ok(close, short_tp, short_sl, bear_tier)

    account  = mt5.account_info()
    balance  = account.balance
    risk_amt = balance * RISK_PERCENT
    long_sz  = calc_lot_size(sym, risk_amt, long_risk)
    short_sz = calc_lot_size(sym, risk_amt, short_risk)

    session          = get_session()
    dst_str          = get_dst_str()
    active_session   = session != "Asian"
    session_threshold = get_session_threshold(session)

    conf_score_bull = calc_confidence_score(
        wk_bull, wk_bear, day_bull, day_bear, htf_bull, htf_bear,
        active_session, ext_bear, s2_long, s2_short,
        any_trig_bull, any_trig_bear, bon_l, bon_s, min_str2, min_bonus, True
    )
    conf_score_bear = calc_confidence_score(
        wk_bull, wk_bear, day_bull, day_bear, htf_bull, htf_bear,
        active_session, ext_bear, s2_long, s2_short,
        any_trig_bull, any_trig_bear, bon_l, bon_s, min_str2, min_bonus, False
    )

    wk_status  = ("BULL" if wk_bull else "BEAR" if wk_bear else "PARTIAL" if wk_partial else "UNCONFIRMED")
    day_status = "BULL" if day_bull else "BEAR" if day_bear else "UNCONFIRMED"

    # ── Print daily P&L status alongside each scan ──
    dd = daily_pnl_state.get("current_pnl", 0)
    print(f"\n{'='*55}")
    print(f"📈 {sym} {tf_name} | {datetime.now(timezone.utc).strftime('%H:%M:%S')} UTC | {session} | {dst_str}")
    print(f"⚡ WkBias:{wk_status} | DayBias:{day_status}")
    print(f"📊 Close:{close:.5f} EMA200:{ema200:.5f} | RSI:{rsi:.1f} ADX:{adx_val:.1f}")
    print(f"💯 ConfBull:{conf_score_bull}/10 ConfBear:{conf_score_bear}/10 | Threshold:{session_threshold}")
    print(f"💰 Daily P&L: {dd:+.2f} | Drawdown limit: {DAILY_DRAWDOWN_LIMIT*100:.0f}%")
    print(f"🚦 Raw — Bull:{raw_bull}({bull_tier}) Bear:{raw_bear}({bear_tier})")

    if bull_sig_final:
        emoji = "🔥" if bull_tier == "STRONG" else "💪" if bull_tier == "MEDIUM" else "⚡"
        if conf_score_bull >= session_threshold:
            print(f"\n{emoji} BUY AUTO — {sym} {tf_name} — {bull_tier} — Score:{conf_score_bull}/10")
            log_trade(sym, tf_name, bull_tier, conf_score_bull, close, long_sl, long_tp, trig_name_bull, sl_src_bull, session, "auto-executed")
            send_telegram(
                f"{emoji} <b>{bull_tier} BUY — {sym} {tf_name}</b>\n\n"
                f"📊 Manny's V4 Mixed Bias | Auto-Executed\n"
                f"💯 Confidence: {conf_score_bull}/10\n"
                f"🎯 Trigger: {trig_name_bull} | SL: {sl_src_bull}\n"
                f"🧭 Weekly: {wk_status} | Daily: {day_status}\n"
                f"📈 Entry: {close:.5f} | SL: {long_sl} | TP: {long_tp}\n"
                f"📦 Lots: {long_sz}\n"
                f"🔔 V4: Partial profit + BE + Trail active\n"
                f"⏰ {datetime.now(timezone.utc).strftime('%H:%M')} UTC"
            )
            place_trade("BUY", long_sl, long_tp, long_sz, sym, tf_name)
            return current_bar, bar_index, last_bear_bar

        elif conf_score_bull >= CONFIDENCE_MANUAL_MIN:
            print(f"\n⚠️ BUY MANUAL — {sym} {tf_name} — Score:{conf_score_bull}/10")
            log_trade(sym, tf_name, bull_tier, conf_score_bull, close, long_sl, long_tp, trig_name_bull, sl_src_bull, session, "manual-alert")
            send_telegram(
                f"⚠️ <b>MANUAL BUY — {sym} {tf_name}</b>\n"
                f"💯 Score: {conf_score_bull}/10 (need {session_threshold})\n"
                f"📈 Entry: {close:.5f} | SL: {long_sl} | TP: {long_tp}"
            )
            return current_bar, bar_index, last_bear_bar

        else:
            print(f"  🔕 BUY ignored — Score:{conf_score_bull}/10")
            log_trade(sym, tf_name, bull_tier, conf_score_bull, close, long_sl, long_tp, trig_name_bull, sl_src_bull, session, "ignored")

    elif bear_sig_final:
        emoji = "🔥" if bear_tier == "STRONG" else "💪" if bear_tier == "MEDIUM" else "⚡"
        if conf_score_bear >= session_threshold:
            print(f"\n{emoji} SELL AUTO — {sym} {tf_name} — {bear_tier} — Score:{conf_score_bear}/10")
            log_trade(sym, tf_name, bear_tier, conf_score_bear, close, short_sl, short_tp, trig_name_bear, sl_src_bear, session, "auto-executed")
            send_telegram(
                f"{emoji} <b>{bear_tier} SELL — {sym} {tf_name}</b>\n\n"
                f"📊 Manny's V4 Mixed Bias | Auto-Executed\n"
                f"💯 Confidence: {conf_score_bear}/10\n"
                f"🎯 Trigger: {trig_name_bear} | SL: {sl_src_bear}\n"
                f"🧭 Weekly: {wk_status} | Daily: {day_status}\n"
                f"📈 Entry: {close:.5f} | SL: {short_sl} | TP: {short_tp}\n"
                f"📦 Lots: {short_sz}\n"
                f"🔔 V4: Partial profit + BE + Trail active\n"
                f"⏰ {datetime.now(timezone.utc).strftime('%H:%M')} UTC"
            )
            place_trade("SELL", short_sl, short_tp, short_sz, sym, tf_name)
            return current_bar, last_bull_bar, bar_index

        elif conf_score_bear >= CONFIDENCE_MANUAL_MIN:
            print(f"\n⚠️ SELL MANUAL — {sym} {tf_name} — Score:{conf_score_bear}/10")
            log_trade(sym, tf_name, bear_tier, conf_score_bear, close, short_sl, short_tp, trig_name_bear, sl_src_bear, session, "manual-alert")
            send_telegram(
                f"⚠️ <b>MANUAL SELL — {sym} {tf_name}</b>\n"
                f"💯 Score: {conf_score_bear}/10 (need {session_threshold})\n"
                f"📈 Entry: {close:.5f} | SL: {short_sl} | TP: {short_tp}"
            )
            return current_bar, last_bull_bar, bar_index

        else:
            print(f"  🔕 SELL ignored — Score:{conf_score_bear}/10")
            log_trade(sym, tf_name, bear_tier, conf_score_bear, close, short_sl, short_tp, trig_name_bear, sl_src_bear, session, "ignored")

    return current_bar, last_bull_bar, last_bear_bar


# ═══════════════════════════════════════════
# TRADE CLOSURE TRACKER
# ─────────────────────────────────────────────
# Tracks which tickets were open in the previous scan loop.
# When a ticket disappears, it means the trade closed — by TP, SL, or early exit.
# We then fetch the deal history from MT5 to get exact close price and P&L.
# ═══════════════════════════════════════════
previously_open_tickets = {}   # {ticket: {symbol, entry, direction, open_time, sl, tp, entry_price}}


def detect_closed_trades() -> None:
    """
    Detect trades that closed since the last scan loop.

    Compares current open positions against previously_open_tickets.
    If a ticket is missing, the trade closed. We then:
      1. Fetch the deal from MT5 history to get exact exit price and P&L
      2. Determine exit reason — TP hit, SL hit, or manual/early exit
      3. Calculate RR achieved and time in trade
      4. Update the Excel journal row for that trade
      5. Send a Telegram notification with the result

    This runs every loop so closures are detected within 60 seconds.
    """
    global previously_open_tickets

    # Get current open positions placed by this bot
    positions     = mt5.positions_get()
    current_ticks = {}

    if positions:
        for pos in positions:
            if pos.magic == 234000:   # only track bot trades
                current_ticks[pos.ticket] = {
                    "symbol":      pos.symbol,
                    "entry":       pos.price_open,
                    "direction":   "BUY" if pos.type == mt5.ORDER_TYPE_BUY else "SELL",
                    "open_time":   datetime.fromtimestamp(pos.time, tz=timezone.utc),
                    "sl":          pos.sl,
                    "tp":          pos.tp,
                }

    # Find tickets that were open before but are now gone
    closed_tickets = {
        t: info for t, info in previously_open_tickets.items()
        if t not in current_ticks
    }

    for ticket, info in closed_tickets.items():
        symbol    = info["symbol"]
        entry     = info["entry"]
        direction = info["direction"]
        open_time = info["open_time"]
        sl        = info["sl"]
        tp        = info["tp"]

        # ── Fetch deal history from MT5 for this ticket ──
        # Look back 24 hours to find the closing deal
        from_time = datetime.now(timezone.utc) - timedelta(hours=24)
        deals     = mt5.history_deals_get(from_time, datetime.now(timezone.utc))

        exit_price  = None
        actual_pnl  = None
        exit_reason = "Unknown"

        if deals:
            # Find the deal that closed this specific position
            # MT5 deals have position_id matching the original ticket
            for deal in deals:
                if deal.position_id == ticket and deal.entry == 1:   # entry==1 means closing deal
                    exit_price = deal.price
                    actual_pnl = deal.profit + deal.commission + deal.swap   # true net P&L
                    break

        # ── Determine exit reason ──
        if exit_price is not None:
            if direction == "BUY":
                # If exit price is near TP — hit take profit
                if tp > 0 and abs(exit_price - tp) < abs(tp - entry) * 0.02:
                    exit_reason = "TP"
                # If exit price is near SL — hit stop loss
                elif sl > 0 and abs(exit_price - sl) < abs(entry - sl) * 0.02:
                    exit_reason = "SL"
                else:
                    exit_reason = "Early-Exit"   # manual or bot-managed exit
            else:  # SELL
                if tp > 0 and abs(exit_price - tp) < abs(entry - tp) * 0.02:
                    exit_reason = "TP"
                elif sl > 0 and abs(exit_price - sl) < abs(sl - entry) * 0.02:
                    exit_reason = "SL"
                else:
                    exit_reason = "Early-Exit"

        # ── Calculate time in trade ──
        mins_open = (datetime.now(timezone.utc) - open_time).total_seconds() / 60

        # ── Calculate RR achieved ──
        risk_val   = abs(entry - sl) if sl > 0 else 1
        rr_achieved = 0.0
        if exit_price is not None and risk_val > 0:
            if direction == "BUY":
                rr_achieved = (exit_price - entry) / risk_val
            else:
                rr_achieved = (entry - exit_price) / risk_val

        # ── Update Excel journal ──
        update_trade_in_journal(
            symbol=symbol,
            timeframe="",          # we match by entry price so TF not needed here
            entry_price=entry,
            exit_reason=exit_reason,
            exit_price=exit_price,
            actual_pnl=actual_pnl,
            time_in_trade_mins=mins_open,
            rr_achieved=rr_achieved,
        )

        # ── Send Telegram notification ──
        if actual_pnl is not None:
            emoji     = "✅" if actual_pnl > 0 else "❌"
            pnl_str   = f"{actual_pnl:+.2f}"
            rr_str    = f"{rr_achieved:.2f}"
            reason_emoji = "🎯" if exit_reason == "TP" else "🛑" if exit_reason == "SL" else "🚪"

            send_telegram(
                f"{emoji} <b>TRADE CLOSED — {symbol}</b>\n\n"
                f"{reason_emoji} Exit Reason: {exit_reason}\n"
                f"📊 Direction: {direction}\n"
                f"📈 Entry: {entry:.5f}\n"
                f"📉 Exit: {exit_price:.5f if exit_price else 'N/A'}\n"
                f"💰 Net P&L: {pnl_str}\n"
                f"📊 RR Achieved: {rr_str}\n"
                f"⏱ Time in trade: {mins_open:.0f} mins\n"
                f"⏰ {datetime.now(timezone.utc).strftime('%H:%M')} UTC"
            )

            print(f"{'✅' if actual_pnl > 0 else '❌'} CLOSED {symbol} #{ticket} | {exit_reason} | P&L:{pnl_str} | RR:{rr_str}")

    # ── Update tracker for next loop ──
    previously_open_tickets = current_ticks


# ═══════════════════════════════════════════
# MAIN LOOP — V4 UPGRADED
# ─────────────────────────────────────────────
# Two key additions vs V3:
# 1. manage_open_trades() called every loop — handles partial, BE, trail, early exits
# 2. check_daily_drawdown() called first — shuts down if limit breached
# ═══════════════════════════════════════════
def run() -> None:
    """
    Main entry point. Connects to MT5 and runs the scanner loop.
    V4 adds: trade management loop, daily drawdown check, Excel journal.
    """
    print("🚀 Manny's Gold Strategy V4 — Tier 1 Upgrades Edition")
    print("=" * 55)
    print("✅ T1-1: Partial Profit Engine (50% close at 1:1)")
    print("✅ T1-2: Break-Even Stop Move (after partial)")
    print("✅ T1-3: Trailing Stop on Runners (swing-based)")
    print("✅ T1-4: Dynamic Exit on Reversal (profit protection)")
    print("✅ T1-5: Early SL Exit (drawdown limiting)")
    print("✅ T1-6: Daily Drawdown Limit (3% shutdown)")
    print("✅ T1-7: Colour-Coded Excel Trade Journal")
    print("=" * 55)

    if not connect_mt5():
        print("Failed to connect to MT5")
        return

    print("✅ MT5 Connected!")

    # Initialise Excel journal on startup
    init_journal()

    h1, h2 = get_pair_bias_hours("XAUUSD")
    dst     = get_dst_str()
    print(f"🕐 {dst} | Gold bias hours: {h1}AM + {h2}AM UTC")

    send_telegram(
        f"🚀 <b>Manny's V4 — Tier 1 Upgrades — Scanner Started</b>\n\n"
        f"✅ Partial profit at 1:1 RR\n"
        f"✅ Break-even stop after partial\n"
        f"✅ Trailing stop on runners\n"
        f"✅ Dynamic early exit on reversal\n"
        f"✅ Early SL exit on aggressive drawdown\n"
        f"✅ Daily drawdown limit: {DAILY_DRAWDOWN_LIMIT*100:.0f}%\n"
        f"✅ Excel journal: trade_journal.xlsx\n"
        f"📈 Pairs: {', '.join(SYMBOLS)}\n"
        f"⏱ Timeframes: 5M 15M 30M 1H 4H\n"
        f"🕐 {dst} | Gold: {h1}AM+{h2}AM UTC"
    )

    # Initialise persistent state
    ob_state     = {}
    signal_state = {}

    for s in SYMBOLS:
        for tf in SYMBOL_TIMEFRAMES.get(s, [mt5.TIMEFRAME_M30]):
            key = f"{s}_{TIMEFRAME_NAMES[tf]}"
            signal_state[key] = {"last_signal_bar": None, "last_bull_bar": None, "last_bear_bar": None}
            ob_state[key]     = {'bull_ob': None, 'bear_ob': None}

    # ── MAIN SCAN LOOP ──
    while True:
        now = datetime.now(timezone.utc)

        # ── T1-6: CHECK DAILY DRAWDOWN FIRST ──
        # If limit breached, skip all trading for this iteration
        if check_daily_drawdown():
            print(f"[{now.strftime('%H:%M:%S')} UTC] 🛑 DAILY DRAWDOWN LIMIT — Bot paused until midnight UTC")
            time.sleep(CHECK_INTERVAL)
            continue   # skip to next loop without scanning or managing trades

        news_black, news_reason = is_news_blackout()

        if is_active_session():
            status = f"🚫 NEWS BLACKOUT: {news_reason}" if news_black else get_session()
            print(f"\n[{now.strftime('%H:%M:%S')} UTC] Scanning | {status}")

            # ── T1-1/2/3/4/5: MANAGE OPEN TRADES FIRST ──
            # Always manage open trades before looking for new signals
            # This ensures we never miss a partial profit or reversal exit
            try:
                manage_open_trades()
            except Exception as e:
                print(f"❌ Trade management error: {e}")

            # ── DETECT CLOSED TRADES ──
            # Check if any previously open trades have now closed
            # Updates journal with exit price, P&L, RR, exit reason
            try:
                detect_closed_trades()
            except Exception as e:
                print(f"❌ Closure detection error: {e}")

            # ── SCAN FOR NEW SIGNALS ──
            if not news_black:
                for sym in SYMBOLS:
                    for tf in SYMBOL_TIMEFRAMES.get(sym, [mt5.TIMEFRAME_M30]):
                        tf_name = TIMEFRAME_NAMES[tf]
                        key     = f"{sym}_{tf_name}"
                        try:
                            state = signal_state[key]
                            new_sig, new_bull, new_bear = check_signal(
                                state["last_signal_bar"],
                                state["last_bull_bar"],
                                state["last_bear_bar"],
                                symbol=sym, timeframe=tf, ob_state=ob_state
                            )
                            signal_state[key]["last_signal_bar"] = new_sig
                            signal_state[key]["last_bull_bar"]   = new_bull
                            signal_state[key]["last_bear_bar"]   = new_bear
                        except Exception as e:
                            print(f"❌ {sym} {tf_name} Error: {e}")
                            send_telegram(f"⚠️ {sym} {tf_name} error: {e}")
        else:
            h1, _ = get_pair_bias_hours("XAUUSD")
            print(f"[{now.strftime('%H:%M:%S')} UTC] 💤 Asian — Next: London 07:00 UTC | Gold Bias: {h1}AM UTC")
            # Still detect closures during Asian session — trades can close overnight
            try:
                detect_closed_trades()
            except Exception as e:
                print(f"❌ Closure detection error (Asian): {e}")

        time.sleep(CHECK_INTERVAL)


if __name__ == "__main__":
    run()
    
    #python scanner_v4.py
    #conda deactivate